import openpyxl
import os

def read_excel_data(sheet_name):
    # Use relative path — works on both Windows and Linux (GitHub Actions)
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(base_dir, "test_data.xlsx")

    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)
    return data

if __name__ == "__main__":
    login_data = read_excel_data("LoginData")
    print("Login Test Data:")
    for row in login_data:
        print(row)
