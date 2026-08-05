import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "LoginData"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2E75B6")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

headers = ["username", "password", "expected_result", "test_type"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

login_data = [
    ("standard_user",   "secret_sauce", "PASS", "positive"),
    ("locked_out_user", "secret_sauce", "FAIL", "negative"),
    ("wrong_user",      "wrong_pass",   "FAIL", "negative"),
    ("standard_user",   "wrong_pass",   "FAIL", "negative"),
    ("",                "",             "FAIL", "negative"),
]

for row, data in enumerate(login_data, 2):
    for col, value in enumerate(data, 1):
        cell = ws1.cell(row=row, column=col, value=value)
        cell.border = border

ws1.column_dimensions["A"].width = 20
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 15

wb.save(r"C:\Users\gpika\Desktop\demo\test_data.xlsx")
print("✅ test_data.xlsx created")
