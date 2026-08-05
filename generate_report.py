import subprocess
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import os

# ── STEP 1: Run pytest and capture output ───────────────────────────────────
print("🚀 Running test suite...")
result = subprocess.run(
    ["pytest", "tests/test_pom.py", "-v", "--tb=short", "--html=report.html"],
    capture_output=True,
    text=True,
    cwd=r"C:\Users\gpika\Desktop\demo"
)
output = result.stdout + result.stderr
print(output)

# ── STEP 2: Parse test results from output ──────────────────────────────────
test_results = []
lines = output.split("\n")

for line in lines:
    # Match lines like: test_saucedemo.py::test_login PASSED
    if "PASSED" in line or "FAILED" in line or "ERROR" in line:
        parts = line.strip().split(" ")
        if "::" in parts[0]:
            full_name = parts[0]
            test_name = full_name.split("::")[-1]
            if "PASSED" in line:
                status = "PASSED"
            elif "FAILED" in line:
                status = "FAILED"
            else:
                status = "ERROR"
            test_results.append((test_name, status))

# ── STEP 3: Map test names to TC IDs and descriptions ───────────────────────
tc_map = {
    "test_login":                        ("TC001", "smoke",      "Valid login with correct credentials"),
    "test_add_to_cart":                  ("TC002", "smoke",      "Add product to cart and verify count"),
    "test_logout":                       ("TC003", "smoke",      "Logout and verify redirect to login page"),
    "test_view_cart":                    ("TC004", "regression", "Open cart and verify correct item is present"),
    "test_checkout_form":                ("TC005", "regression", "Fill checkout form and proceed to overview"),
    "test_place_order":                  ("TC006", "regression", "Place order and verify confirmation message"),
    "test_invalid_login_wrong_password": ("TC007", "negative",   "Login with wrong password shows error"),
    "test_invalid_login_empty_fields":   ("TC008", "negative",   "Login with empty fields shows error"),
    "test_locked_out_user":              ("TC009", "negative",   "Locked out user is blocked from login"),
}

# ── STEP 4: Create Excel Workbook ───────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Execution Report"

# Colors
DARK_BLUE   = "1F3864"
LIGHT_BLUE  = "BDD7EE"
GREEN       = "C6EFCE"
RED         = "FFC7CE"
YELLOW      = "FFEB9C"
WHITE       = "FFFFFF"
GREY        = "F2F2F2"

# Fonts
header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
title_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
normal_font  = Font(name="Calibri", size=10)
pass_font    = Font(name="Calibri", bold=True, color="375623", size=10)
fail_font    = Font(name="Calibri", bold=True, color="9C0006", size=10)

# Border
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Title Row ───────────────────────────────────────────────────────────────
ws.merge_cells("A1:G1")
ws["A1"] = "🧪 AUTOMATION TEST EXECUTION REPORT — SauceDemo Web App"
ws["A1"].font = title_font
ws["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# ── Metadata Row ────────────────────────────────────────────────────────────
ws.merge_cells("A2:G2")
run_time = datetime.now().strftime("%d %B %Y  |  %I:%M %p")
ws["A2"] = f"Executed By: Gopika Sundar     |     Date & Time: {run_time}     |     Environment: https://www.saucedemo.com"
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
ws["A2"].fill = PatternFill("solid", fgColor="2F5496")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# ── Empty Row ───────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 8

# ── Column Headers ──────────────────────────────────────────────────────────
headers = ["TC ID", "Test Case Name", "Description", "Type", "Status", "Remarks", "Executed On"]
ws.append([""] * 7)  # row 3 spacer
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num, value=header)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[4].height = 22

# ── Data Rows ────────────────────────────────────────────────────────────────
row_num = 5
passed_count = 0
failed_count = 0

for test_name, status in test_results:
    tc_id, tc_type, description = tc_map.get(test_name, ("N/A", "N/A", test_name))

    remarks = "Test executed successfully" if status == "PASSED" else "Test failed — check logs"
    executed_on = datetime.now().strftime("%d-%b-%Y %I:%M %p")

    row_data = [tc_id, test_name, description, tc_type.upper(), status, remarks, executed_on]

    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.font = normal_font
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Row background alternating
        bg = WHITE if row_num % 2 == 0 else GREY
        cell.fill = PatternFill("solid", fgColor=bg)

        # Status cell color
        if col_num == 5:
            if status == "PASSED":
                cell.fill = PatternFill("solid", fgColor=GREEN)
                cell.font = pass_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                passed_count += 1
            elif status == "FAILED":
                cell.fill = PatternFill("solid", fgColor=RED)
                cell.font = fail_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                failed_count += 1

        # Type cell color
        if col_num == 4:
            if value == "SMOKE":
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            elif value == "REGRESSION":
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            elif value == "NEGATIVE":
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row_num].height = 20
    row_num += 1

# ── Summary Section ──────────────────────────────────────────────────────────
row_num += 1
total = passed_count + failed_count

summary_data = [
    ("Total Test Cases", total),
    ("Passed", passed_count),
    ("Failed", failed_count),
    ("Pass Rate", f"{int((passed_count/total)*100) if total > 0 else 0}%"),
]

ws.cell(row=row_num, column=1, value="📊 SUMMARY").font = Font(bold=True, size=11, color=DARK_BLUE)
row_num += 1

for label, value in summary_data:
    label_cell = ws.cell(row=row_num, column=1, value=label)
    value_cell = ws.cell(row=row_num, column=2, value=value)
    label_cell.font = Font(bold=True, size=10)
    value_cell.font = Font(bold=True, size=10,
                           color="375623" if label == "Passed"
                           else "9C0006" if label == "Failed"
                           else "000000")
    label_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    value_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    label_cell.border = border
    value_cell.border = border
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_num].height = 18
    row_num += 1

# ── Column Widths ────────────────────────────────────────────────────────────
col_widths = [10, 38, 48, 14, 12, 35, 22]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ── Save File ────────────────────────────────────────────────────────────────
output_path = r"C:\Users\gpika\Desktop\demo\TestExecutionReport.xlsx"
wb.save(output_path)
print(f"\n✅ Excel report saved → {output_path}")
print(f"📊 Summary: {passed_count} Passed | {failed_count} Failed | {total} Total")

# ── Auto Open the Excel file ─────────────────────────────────────────────────
os.startfile(output_path)
